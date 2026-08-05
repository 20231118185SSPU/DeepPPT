#!/usr/bin/env python3
"""Regenerate the OfficeCLI XLSX inspection fixture.

Synthetic workbook with two sheets, formulas, one Excel table, a named range,
and a data validation. Deterministic output (fixed zip timestamps; the default
font is reordered to the Excel order OfficeCLI's schema validator expects).
No user or restricted content.

Run from the repo root:
    python skills/ppt-master/fixtures/officecli/rebuild_xlsx_source.py
"""

import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "xlsx_source.xlsx"


def _finalize(path: Path) -> None:
    """Reorder the default font (Excel order) and rezip deterministically."""
    with zipfile.ZipFile(path) as z:
        parts = {i.filename: z.read(i.filename) for i in z.infolist()}
    xml = parts["xl/styles.xml"].decode("utf-8")
    xml = re.sub(
        r"<font>.*?</font>",
        '<font><sz val="11"/><color theme="1"/><name val="Calibri"/>'
        '<family val="2"/><scheme val="minor"/></font>',
        xml, count=1, flags=re.S,
    )
    parts["xl/styles.xml"] = xml.encode("utf-8")
    # Neutralize openpyxl's core.xml timestamps.
    core = parts.get("docProps/core.xml", b"")
    if core:
        core_text = core.decode("utf-8")
        core_text = re.sub(
            r"(?<=\>)\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}Z(?=</dcterms:(?:created|modified)>)",
            "1980-01-01T00:00:00Z", core_text,
        )
        parts["docProps/core.xml"] = core_text.encode("utf-8")
    tmp = path.with_name(path.name + ".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for name in sorted(parts):
            z.writestr(zipfile.ZipInfo(name, (1980, 1, 1, 0, 0, 0)), parts[name])
    path.unlink()
    tmp.replace(path)


def main() -> int:
    wb = Workbook()

    # Sheet 1: Sales data + formula + table + validation + named range
    ws = wb.active
    ws.title = "Sales"
    headers = ["Product", "Q1", "Q2", "Q3", "Total"]
    ws.append(headers)
    rows = [
        ["Widget A", 120, 135, 148, None],
        ["Widget B", 89, 92, 97, None],
        ["Gadget X", 45, 52, 61, None],
    ]
    for row in rows:
        ws.append(row)
    for r in range(2, 5):
        ws.cell(row=r, column=5).value = f"=SUM(B{r}:D{r})"
    ws.append(["Totals", "=SUM(B2:B4)", "=SUM(C2:C4)", "=SUM(D2:D4)", "=SUM(E2:E4)"])

    tab = Table(displayName="SalesTable", ref="A1:E5")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)

    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="1000")
    dv.add("B2:D4")
    ws.add_data_validation(dv)

    wb.defined_names.add(DefinedName("SalesRange", attr_text="Sales!$A$1:$E$5"))

    # Sheet 2: Notes
    ws2 = wb.create_sheet("Notes")
    ws2.append(["Note", "Value"])
    ws2.append(["Alpha", 1])
    ws2.append(["Beta", 2])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    _finalize(OUT)
    print(f"Written {OUT}")
    print("  Sheets: 2, formulas: 7, table: 1, validation: 1, named range: 1")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
